#!/usr/bin/env python3
"""
Run ALL 669 queries against Azure and OCI RAG APIs.
Saves detailed Excel report with per-query comparison.

Usage:
    python test_all_queries.py
"""

import json
import re
import csv
import time
import requests
import sys
from pathlib import Path
from datetime import datetime

AZURE = "http://localhost:7870"
OCI = "http://localhost:7874"
TIMEOUT = 30
REPORT_JSON = "/tmp/full_comparison_report.json"
REPORT_EXCEL = "/tmp/oci_vs_azure_comparison.xlsx"


def strip_html(t):
    return re.sub(r'<[^>]+>', '', t).strip()


def get_all_queries():
    queries = set()
    for logfile in ['/home/admincsp/graphiti_fixed_test/logs/llm_chunked_server.log',
                    '/home/admincsp/graphiti_fixed_test/logs/rag_server.log',
                    '/home/admincsp/graphiti_fixed_test/logs/tool_call_server.log']:
        try:
            with open(logfile) as f:
                for line in f:
                    m = re.search(r'"query":\s*"([^"]+)"', line)
                    if m:
                        q = m.group(1).replace('\\n', ' ').strip()
                        if len(q) > 5:
                            queries.add(q)
        except:
            pass
    try:
        with open('/home/admincsp/frontend_integration/CSP BrainShift GenAI Chatbot Test Template.xlsx - Chatbot Test Questions (1).csv') as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                if len(row) >= 2 and row[1].strip() and len(row[1].strip()) > 5:
                    queries.add(row[1].strip())
    except:
        pass

    skip = {'hello', 'hi', 'hey', 'thanks', 'thank you', 'ok', 'okay', 'yes', 'no', 'bye', 'sure', 'great'}
    return sorted([q for q in queries if q.lower().strip() not in skip and len(q.split()) >= 2])


def query_api(url, query, user_id):
    t0 = time.time()
    try:
        r = requests.post(f"{url}/query", json={"query": query, "user_id": user_id}, timeout=TIMEOUT)
        elapsed = time.time() - t0
        d = r.json()
        resp = strip_html(d.get("response", ""))
        return {
            "status": "ok",
            "elapsed": round(elapsed, 2),
            "route": d.get("metadata", {}).get("route", "?"),
            "sources": len(d.get("metadata", {}).get("sources", [])),
            "response": resp,
            "response_length": len(resp),
        }
    except requests.exceptions.Timeout:
        return {"status": "timeout", "elapsed": round(time.time() - t0, 2), "response": "", "response_length": 0, "route": "?", "sources": 0}
    except Exception as e:
        return {"status": "error", "elapsed": round(time.time() - t0, 2), "response": "", "response_length": 0, "route": "?", "sources": 0, "error": str(e)[:100]}


def determine_winner(az, oci_r):
    az_ok = az["status"] == "ok" and az["response_length"] > 20
    oci_ok = oci_r["status"] == "ok" and oci_r["response_length"] > 20

    if az_ok and not oci_ok:
        return "azure"
    if oci_ok and not az_ok:
        return "oci"
    if not az_ok and not oci_ok:
        return "both_failed"

    # Both answered
    az_clarify = any(kw in az["response"].lower() for kw in
                     ["are you asking about", "which country", "please specify", "could you clarify", "could you please specify"])
    oci_clarify = any(kw in oci_r["response"].lower() for kw in
                      ["are you asking about", "which country", "please specify", "could you clarify", "could you please specify"])

    if az_clarify and not oci_clarify:
        return "oci"  # OCI gave direct answer
    if oci_clarify and not az_clarify:
        return "oci"  # OCI smartly asked clarification
    return "tie"


def main():
    queries = get_all_queries()
    print(f"Testing {len(queries)} queries")
    print(f"Azure: {AZURE}")
    print(f"OCI:   {OCI}")
    print(f"Started: {datetime.now().strftime('%H:%M:%S')}")
    print()

    report = []
    stats = {"azure": 0, "oci": 0, "tie": 0, "both_failed": 0}
    az_total_time = oci_total_time = 0

    for i, q in enumerate(queries, 1):
        sys.stdout.write(f"\r[{i:3d}/{len(queries)}] ")
        sys.stdout.flush()

        az = query_api(AZURE, q, f"az_{i}")
        oci_r = query_api(OCI, q, f"oci_{i}")
        winner = determine_winner(az, oci_r)

        stats[winner] += 1
        az_total_time += az["elapsed"]
        oci_total_time += oci_r["elapsed"]

        report.append({
            "index": i,
            "question": q,
            "winner": winner,
            "azure_time": az["elapsed"],
            "azure_route": az["route"],
            "azure_sources": az["sources"],
            "azure_length": az["response_length"],
            "azure_response": az["response"][:1000],
            "azure_status": az["status"],
            "oci_time": oci_r["elapsed"],
            "oci_route": oci_r["route"],
            "oci_sources": oci_r["sources"],
            "oci_length": oci_r["response_length"],
            "oci_response": oci_r["response"][:1000],
            "oci_status": oci_r["status"],
        })

        if i % 50 == 0:
            print(f"\n  [{i}/{len(queries)}] Az={stats['azure']} OCI={stats['oci']} Tie={stats['tie']} Fail={stats['both_failed']}")

    # Save JSON
    with open(REPORT_JSON, "w") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)

    # Save Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["OCI vs Azure RAG Comparison Report"])
        ws_summary.append([f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        ws_summary.append([f"Total queries: {len(queries)}"])
        ws_summary.append([])
        ws_summary.append(["Metric", "Count", "Percentage"])
        ws_summary.append(["Azure wins", stats["azure"], f"{stats['azure']*100//len(queries)}%"])
        ws_summary.append(["OCI wins", stats["oci"], f"{stats['oci']*100//len(queries)}%"])
        ws_summary.append(["Ties", stats["tie"], f"{stats['tie']*100//len(queries)}%"])
        ws_summary.append(["Both failed", stats["both_failed"], f"{stats['both_failed']*100//len(queries)}%"])
        ws_summary.append([])
        ws_summary.append(["Avg latency Azure", f"{az_total_time/len(queries):.2f}s"])
        ws_summary.append(["Avg latency OCI", f"{oci_total_time/len(queries):.2f}s"])

        # Bold headers
        for cell in ws_summary[1]:
            cell.font = Font(bold=True, size=14)
        for cell in ws_summary[5]:
            cell.font = Font(bold=True)

        # Detail sheet
        ws = wb.create_sheet("Detailed Results")
        headers = ["#", "Question", "Winner", "Azure Time", "Azure Route", "Azure Sources",
                    "Azure Response Length", "Azure Response", "OCI Time", "OCI Route",
                    "OCI Sources", "OCI Response Length", "OCI Response"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font

        # Color fills
        azure_fill = PatternFill(start_color="DAEEF3", fill_type="solid")
        oci_fill = PatternFill(start_color="E2EFDA", fill_type="solid")
        tie_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
        fail_fill = PatternFill(start_color="FCE4EC", fill_type="solid")

        for e in report:
            row = [
                e["index"], e["question"], e["winner"],
                e["azure_time"], e["azure_route"], e["azure_sources"],
                e["azure_length"], e["azure_response"][:500],
                e["oci_time"], e["oci_route"], e["oci_sources"],
                e["oci_length"], e["oci_response"][:500],
            ]
            ws.append(row)

            # Color the row based on winner
            row_num = ws.max_row
            fill = {"azure": azure_fill, "oci": oci_fill, "tie": tie_fill, "both_failed": fail_fill}.get(e["winner"], tie_fill)
            for col in range(1, 14):
                ws.cell(row=row_num, column=col).fill = fill

        # Column widths
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['H'].width = 80
        ws.column_dimensions['M'].width = 80
        ws.column_dimensions['C'].width = 10

        wb.save(REPORT_EXCEL)
        print(f"\n\nExcel saved: {REPORT_EXCEL}")
    except ImportError:
        print("\n\nopenpyxl not installed — saving CSV instead")
        csv_path = REPORT_EXCEL.replace('.xlsx', '.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=report[0].keys())
            writer.writeheader()
            writer.writerows(report)
        print(f"CSV saved: {csv_path}")

    # Print summary
    total = len(queries)
    print(f"\n{'='*60}")
    print(f"FINAL RESULTS: {total} queries")
    print(f"{'='*60}")
    print(f"Azure wins:  {stats['azure']:4d} ({stats['azure']*100//total}%)")
    print(f"OCI wins:    {stats['oci']:4d} ({stats['oci']*100//total}%)")
    print(f"Ties:        {stats['tie']:4d} ({stats['tie']*100//total}%)")
    print(f"Both failed: {stats['both_failed']:4d} ({stats['both_failed']*100//total}%)")
    print(f"Avg latency: Azure={az_total_time/total:.2f}s  OCI={oci_total_time/total:.2f}s")
    print(f"\nJSON: {REPORT_JSON}")
    print(f"Excel: {REPORT_EXCEL}")


if __name__ == "__main__":
    main()
